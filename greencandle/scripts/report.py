#!/usr/bin/env python
#pylint: disable=wrong-import-position

"""
Create Excel Spreadsheet with results and analysis of trades
"""

import sys
import openpyxl
from greencandle.lib import config
config.create_config()
from greencandle.lib.mysql import Mysql

def main():
    """
    main function
    """

    if len(sys.argv) > 1 and sys.argv[1] == '--help':
        print("Generate Excel report from database entries")
        sys.exit(0)

    if len(sys.argv) != 3:
        sys.stderr.write("Usage: report <interval> <filename>\n")
        sys.exit(1)

    interval = sys.argv[1]
    filename = sys.argv[2]
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))

    mysql = Mysql(test=True, interval=interval)
    queries = {"weekly": "select perc, pair, week(sell_time) as week from profit",
               "monthly": "select perc, pair, month(sell_time) as month from profit",
               "average-day": "select pair, hour(timediff(sell_time,buy_time)) as hours from \
                       trades where sell_time != '0000-00-00 00:00:00' group by pair;",
               "perc-month": "select pair, sum(perc)  as perc, month(sell_time) as month from \
                        profit where sell_time != '0000-00-00 00:00:00'  group by pair, month \
                        order by month,perc;",
               "profit-pair": "select pair, sum(perc) as perc from profit where \
                        sell_time != '0000-00-00 00:00:00' group by pair;",
               "trades": "select buy_time, sell_time, buy_price, sell_price, base_profit, \
                          perc, drawdown_perc from profit;",
               "hours-pair": "select pair, sum(hour(timediff(sell_time,buy_time))) \
                        as hours from trades where sell_time \
                        != '0000-00-00 00:00:00' group by pair",
               "profit-factor": "select IFNULL((select sum(base_profit) from profit where \
                                base_profit >0)/-(select sum(base_profit) from profit where \
                                base_profit <0),100) as profit_factor",
               "buy-hold-return": "select (select buy_price from profit order by buy_time limit 1) \
                                   as first_buy, (select sell_price from profit order by \
                                   buy_time desc limit 1) as last_sell, (select \
                                   (last_sell-first_buy)/first_buy)*100 as buy_hold"}
    for name, query in queries.items():
        result = mysql.fetch_sql_data(query)
        workbook.create_sheet(title=name)
        sheet = workbook.get_sheet_by_name(name)
        for row_no, row in enumerate(result):
            for col_no, col in enumerate(row):
                sheet.cell(row=row_no+1, column=col_no+1).value = col
        workbook.save(filename)

if __name__ == "__main__":
    main()
