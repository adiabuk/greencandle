#!/usr/bin/env python
#pylint: disable=wrong-import-position,no-member,too-many-locals

"""
Create Excel Spreadsheet with results and analysis of trades
"""

import sys
import openpyxl
from greencandle.lib import config
from greencandle.lib.balance_common import get_quote
config.create_config()
from greencandle.lib.mysql import Mysql

def main():
    """
    main function

    Creates Excel spreadsheet from trades tables with the following columns:
    ExDate
    from token
    from token amount
    exchange to token
    to token amount
    GBP Amount

    Each database row accounts for 2 transactions as there is an open and close

    """

    if len(sys.argv) > 1 and sys.argv[1] == '--help':
        print("Generate Excel accounting report from database entries")
        sys.exit(0)

    if len(sys.argv) != 2:
        sys.stderr.write("Usage: accounts <filename>\n")
        sys.exit(1)

    filename = sys.argv[1]
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))
    query = """Select open_time, close_time, pair, quote_in, quote_out, base_in, base_out,
               open_usd_rate, close_usd_rate, open_gbp_rate, close_gbp_rate from trades;"""
    mysql = Mysql(test=True)
    results = mysql.fetch_sql_data(query, header=False)
    agg = [("time", "from token", "from token amount", "exchange to token",
            "to token amount", "GBP Amount")]

    for result in results:
        # Get fields for current row
        open_time, close_time, pair, quote_in, quote_out, base_in, base_out, \
                open_usd_rate, close_usd_rate, open_gbp_rate, close_gbp_rate = result

        # open trade
        from_token = get_quote(pair)
        exchange_token = pair.rstrip(from_token)
        try:
            gbp_amount = float(base_in) * float(open_usd_rate) * float(open_gbp_rate)
        except TypeError:
            gbp_amount = ""
        agg.append((open_time, from_token, base_in, exchange_token,
                    quote_in, gbp_amount))

        # close trade
        try:
            gbp_amount = float(base_out) * float(close_usd_rate) * float(close_gbp_rate)
        except TypeError:
            gbp_amount = ""
        agg.append((close_time, exchange_token, quote_out, from_token, base_out, gbp_amount))

    sheet = workbook.create_sheet(title="Orderbook")
    for row_no, row in enumerate(agg):
        for col_no, _ in enumerate(row):
            sheet.cell(column=col_no+1, row=row_no+1, value=agg[row_no][col_no])

    workbook.save(filename)

if __name__ == "__main__":
    main()
