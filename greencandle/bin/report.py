#!/usr/bin/env python
#pylint: disable=wrong-import-position,no-member

"""
Create Excel Spreadsheet with results and analysis of trades
"""

import sys
import openpyxl
from greencandle.lib.common import arg_decorator
from greencandle.lib import config
config.create_config()
from greencandle.lib.mysql import Mysql

@arg_decorator
def main():
    """
    Create Excel Spreadsheet with results and analysis of trades
    Analysis comes from database entries: open/close prices as
    well as drawup/drawdown prices and profits

    Usage: report <interval> <filename>
    """

    if len(sys.argv) != 3:
        sys.stderr.write("Usage: report <interval> <filename>\n")
        sys.exit(1)

    interval = sys.argv[1]
    filename = sys.argv[2]
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))
    if config.main.trade_direction == 'short':
        open_hold = "(select (first_open-last_close)/first_open)*100 as open_hold"
    elif config.main.trade_direction == 'long':
        open_hold = "(select (last_close-first_open)/first_open)*100 as open_hold"

    mysql = Mysql(test=True, interval=interval)
    queries = {"weekly": "select perc, pair, week(close_time) as week from profit",
               "monthly": "select perc, pair, month(close_time) as month from profit",
               "average-day": "select pair, hour(timediff(close_time,open_time)) as hours from \
                       trades where close_time != '0000-00-00 00:00:00' group by pair;",
               "perc-month": "select pair, sum(perc)  as perc, month(close_time) as month from \
                        profit where close_time != '0000-00-00 00:00:00'  group by pair, month \
                        order by month,perc;",
               "profit-pair": "select pair, sum(perc) as perc from profit where \
                        close_time != '0000-00-00 00:00:00' group by pair;",
               "trades": "select open_time, close_time, open_price, close_price, quote_profit, \
                          perc, drawdown_perc from profit;",
               "hours-pair": "select pair, sum(hour(timediff(close_time,open_time))) \
                        as hours from trades where close_time \
                        != '0000-00-00 00:00:00' group by pair",
               "profit-factor": "select IFNULL((select sum(quote_profit) from profit where \
                                quote_profit >0)/-(select sum(quote_profit) from profit where \
                                quote_profit <0),100) as profit_factor",
               "open-hold-return": "select (select open_price from profit order by \
                                   open_time limit 1) \
                                   as first_open, (select close_price from profit order by \
                                   open_time desc limit 1) as last_close," + open_hold}
    for name, query in queries.items():
        result = mysql.fetch_sql_data(query)
        workbook.create_sheet(title=name)
        sheet = workbook.get_sheet_by_name(name)
        for row_no, row in enumerate(result):
            for col_no, col in enumerate(row):
                sheet.cell(row=row_no+1, column=col_no+1).value = col
        workbook.save(filename)

if __name__ == "__main__":
    main()
